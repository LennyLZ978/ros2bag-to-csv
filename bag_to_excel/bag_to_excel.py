#!/usr/bin/env python3
"""
ROS2 subscriber node that records /mpc_traj and /odom into CSV files
(CSV opens natively in Excel — no third-party libraries needed).

Workflow
--------
Terminal 1 — start this recorder:
    python3 bag_to_excel.py [--namespace <ns>] [--output-dir <dir>]

Terminal 2 — play the bag:
    ros2 bag play <bag_path>

Press Ctrl+C in terminal 1 when the bag finishes. Two CSV files are written:
    <output_dir>/mpc_traj.csv
    <output_dir>/odom.csv

Topic resolution
----------------
No namespace (default):   /mpc_traj   /odom
With --namespace robot1:  /robot1/mpc_traj   /robot1/odom

Output layout
-------------
mpc_traj.csv  (timestamps as row groups, poses as columns, alternate messages only):
    timestamp , Poses[0] , Poses[1] , ...  , Diff timestamp
    t1        , pos.x[0] , pos.x[1] , ...  , dt    <- x row
              , pos.y[0] , pos.y[1] , ...  ,       <- y row
              , pos.z[0] , pos.z[1] , ...  ,       <- z row
    t3        , ...
    (only alternate timestamps: t1, t3, t5, …)

odom.csv  (vertical, alternate timestamps only):
    timestamp , Poses
    t1        , pos.x    <- x row
              , pos.y    <- y row
              , pos.z    <- z row
    t3        , ...
"""

import argparse
import csv
import math
import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import rclpy
from rclpy.node import Node
from rclpy.qos import QoSProfile, ReliabilityPolicy, DurabilityPolicy
from nav_msgs.msg import Odometry, Path


# ── helper ────────────────────────────────────────────────────────────────────

def yaw_from_quaternion(x: float, y: float, z: float, w: float) -> float:
    return math.atan2(2.0 * (w * z + x * y), 1.0 - 2.0 * (y * y + z * z))


# ── ROS2 node ─────────────────────────────────────────────────────────────────

class BagRecorder(Node):
    def __init__(self, traj_topic: str, odom_topic: str, output_dir: str):
        super().__init__("bag_recorder")

        self._output_dir = output_dir
        self._mpc_data: list[tuple[float, Path]] = []      # (t_sec, msg)
        self._odom_data: list[tuple[float, Odometry]] = []

        qos = QoSProfile(
            depth=10,
            reliability=ReliabilityPolicy.BEST_EFFORT,
            durability=DurabilityPolicy.VOLATILE,
        )

        self.create_subscription(Path,     traj_topic, self._mpc_cb,  qos)
        self.create_subscription(Odometry, odom_topic, self._odom_cb, qos)

        self.get_logger().info(f"Subscribed to  {traj_topic}  and  {odom_topic}")
        self.get_logger().info("Play your bag now. Press Ctrl+C when done.")

    # ── callbacks ──────────────────────────────────────────────────────────────

    def _mpc_cb(self, msg: Path):
        t = self._ros_time_to_sec(msg.header.stamp)
        self._mpc_data.append((t, msg))
        self.get_logger().debug(
            f"/mpc_traj  t={t:.3f}  poses={len(msg.poses)}", throttle_duration_sec=1.0
        )

    def _odom_cb(self, msg: Odometry):
        t = self._ros_time_to_sec(msg.header.stamp)
        self._odom_data.append((t, msg))
        self.get_logger().debug(
            f"/odom  t={t:.3f}", throttle_duration_sec=1.0
        )

    # ── write ──────────────────────────────────────────────────────────────────

    def write_csv(self):
        os.makedirs(self._output_dir, exist_ok=True)
        traj_path = os.path.join(self._output_dir, "mpc_traj.csv")
        odom_path = os.path.join(self._output_dir, "odom.csv")

        self._write_mpc_traj_csv(traj_path)
        self._write_odom_csv(odom_path)

        # Apply diagonal highlighting and save as .xlsx
        traj_xlsx = self._highlight_diagonal_xlsx(traj_path)
        odom_xlsx = self._highlight_diagonal_xlsx(odom_path)

        self.get_logger().info(
            f"\nSaved:\n  {traj_path}  ({len(self._mpc_data)} messages)\n"
            f"  {odom_path}  ({len(self._odom_data)} messages)\n"
            f"  {traj_xlsx}  (with diagonal highlighting)\n"
            f"  {odom_xlsx}  (with diagonal highlighting)"
        )

    def _write_mpc_traj_csv(self, path: str):
        """
        timestamp , Poses[0] , Poses[1] , ...  , Diff timestamp
        t1        , pos.x[0] , pos.x[1] , ...  , dt             <- x row
                  , pos.y[0] , pos.y[1] , ...  ,                <- y row
                  , pos.z[0] , pos.z[1] , ...  ,                <- z row
        t3        , ...
        (only alternate messages are recorded: t1, t3, t5, …)
        """
        if not self._mpc_data:
            print("[mpc_traj] No messages received — CSV not written.")
            return

        # Keep only alternate messages (index 0, 2, 4, …)
        alt_data = self._mpc_data[::2]

        max_poses = max(len(msg.poses) for _, msg in alt_data)

        with open(path, "w", newline="") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(
                ["timestamp"]
                + [f"Poses[{i}]" for i in range(max_poses)]
                + ["Diff timestamp"]
            )

            # One group of 3 rows per message
            prev_t = None
            for t_sec, msg in alt_data:
                rows = [[] for _ in range(3)]  # x, y, z sub-rows

                for ps in msg.poses:
                    p = ps.pose.position
                    rows[0].append(round(p.x, 9))
                    rows[1].append(round(p.y, 9))
                    rows[2].append(round(p.z, 9))

                # Pad shorter pose lists with empty strings
                for r in rows:
                    while len(r) < max_poses:
                        r.append("")

                # Compute timestamp difference from previous trajectory
                if prev_t is None:
                    dt = 0.0
                else:
                    dt = (t_sec - prev_t) * 1000.0
                prev_t = t_sec

                writer.writerow([round(t_sec, 9)] + rows[0] + [round(dt, 9)])  # x
                writer.writerow([""]              + rows[1] + [""])             # y
                writer.writerow([""]              + rows[2] + [""])             # z

    def _write_odom_csv(self, path: str):
        """
        timestamp , Poses
        t1        , pos.x   <- x row
                  , pos.y   <- y row
                  , pos.z   <- z row
        t3        , ...
        (alternate timestamps only)
        """
        if not self._odom_data:
            print("[odom] No messages received — CSV not written.")
            return

        # Keep only alternate messages (index 0, 2, 4, …)
        alt_data = self._odom_data[::2]

        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["timestamp", "Poses"])

            for t_sec, msg in alt_data:
                p = msg.pose.pose.position
                writer.writerow([round(t_sec, 9), round(p.x, 9)])  # x
                writer.writerow(["",              round(p.y, 9)])   # y
                writer.writerow(["",              round(p.z, 9)])   # z

    @staticmethod
    def _highlight_diagonal_xlsx(csv_path: str, rows_per_group: int = 3):
        """Read a CSV, apply diagonal highlighting with 5 alternating colours,
        and save alongside as .xlsx."""
        xlsx_path = csv_path.rsplit(".", 1)[0] + ".xlsx"

        # Load CSV into a workbook via openpyxl (re-read so we don't duplicate
        # the writing logic — keeps the change minimal).
        import csv as _csv
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        with open(csv_path, newline="") as f:
            for row in _csv.reader(f):
                ws.append(row)

        colours = [
            PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # light blue
            PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # light green
            PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),  # lemon chiffon
            PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid"),  # light orange
            PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # light pink
        ]

        # Number of data columns (Poses columns only, exclude timestamp & diff)
        num_pose_cols = ws.max_column - 2  # first col = timestamp, last = diff
        if num_pose_cols <= 0:
            wb.save(xlsx_path)
            return xlsx_path

        # Row 1 is header; data groups start at row 2
        num_groups = (ws.max_row - 1) // rows_per_group

        num_colours = len(colours)
        for g in range(num_groups):
            for c in range(num_pose_cols):
                col_idx = c + 2  # +2 because col 1 = timestamp
                fill = colours[(g + c) % num_colours]
                for r_offset in range(rows_per_group):
                    row_num = 2 + g * rows_per_group + r_offset
                    ws.cell(row=row_num, column=col_idx).fill = fill

        wb.save(xlsx_path)
        return xlsx_path

    # ── util ───────────────────────────────────────────────────────────────────

    @staticmethod
    def _ros_time_to_sec(stamp) -> float:
        return stamp.sec + stamp.nanosec * 1e-9


# ── entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Record /mpc_traj and /odom to CSV while ros2 bag play runs."
    )
    parser.add_argument(
        "--namespace", "-n",
        default="",
        help="Optional topic namespace (e.g. 'robot1' → /robot1/mpc_traj, /robot1/odom). "
             "Leave empty for no namespace.",
    )
    parser.add_argument(
        "--output-dir", "-o",
        default=os.path.dirname(os.path.abspath(__file__)),
        help="Directory to write mpc_traj.csv and odom.csv (default: same directory as this script).",
    )

    # argparse and rclpy both consume sys.argv; split them cleanly
    args, ros_args = parser.parse_known_args()

    ns = args.namespace.strip("/")
    prefix = f"/{ns}" if ns else ""
    traj_topic = f"{prefix}/mpc_traj"
    odom_topic = f"{prefix}/odom"

    rclpy.init(args=ros_args or None)
    node = BagRecorder(traj_topic, odom_topic, args.output_dir)

    try:
        while rclpy.ok():
            rclpy.spin_once(node, timeout_sec=0.1)
    except KeyboardInterrupt:
        pass
    finally:
        print("\nShutting down — writing CSV files...")
        node.write_csv()
        node.destroy_node()
        rclpy.try_shutdown()


if __name__ == "__main__":
    main()
